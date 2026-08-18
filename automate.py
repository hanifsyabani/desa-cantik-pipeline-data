"""
generate_dashboard.py
----------------------
Otomasi transformasi sheet "3.list_data (value)" (format WIDE, 1 baris per
responden) menjadi sheet "4.template dashboard" (format LONG, 1 baris per
responden x komoditas) — menggantikan proses copy-paste 20 blok formula manual.

CARA PAKAI:
    python3 generate_dashboard.py <file_input.xlsx> <file_output.xlsx>

Contoh:
    python3 generate_dashboard.py fix_laporan_kuisioner_clean.xlsx hasil_update.xlsx

Setiap kali data responden di sheet 3 bertambah/berubah, cukup jalankan ulang
script ini terhadap file terbaru -- sheet 4 akan di-generate ulang otomatis
tanpa perlu tarik-tarik formula.
"""

import sys
import re
import datetime
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

SRC_SHEET = "3.list_data (value)"
QUALITY_SHEET = "Kode Kualitas"
DASHBOARD_SHEET = "4.template dashboard"

# Baris pertama data (header ada di baris 4) di sheet 3
SRC_HEADER_ROW = 4
SRC_DATA_START_ROW = 5

# Kolom identitas wilayah di sheet 3 -> kolom A-E di sheet 4
ID_COLS = ["O", "P", "Q", "R", "S"]

# Konfigurasi 20 blok komoditas, hasil ekstraksi dari formula asli sheet 4.
# kualitas_col=None  -> kualitas SELALU tetap (fixed_h / fixed_i), tidak
#                        tergantung jawaban responden (komoditas dgn 1 pilihan kualitas saja)
# kualitas_col diisi -> id_kualitas ditentukan dari digit pertama jawaban
#                        (mis. "1. Selancar: Per-Kg" -> pilihan 1), lalu di-mapping
#                        lewat choice_map
#
# CATATAN PERBAIKAN BUG: pada template asli, kolom "Asal Barang" untuk
# Gula Pasir salah merujuk ke kolom Asal Barang milik Minyak Goreng (AI,
# harusnya AM). Sudah diperbaiki di bawah ini (asal_col='AM').
BLOCKS = [
    dict(id_kom=1,  nama="Beras",              unit="kg",
         kualitas_col="X",  harga_col="Y",  tanggal_col="Z",  asal_col="AA",
         choice_map={1: 1, 2: 3, 3: 18},
         choice_names={1: "Selancar", 2: "Topi Koki", 3: "Lainnya"}),
    dict(id_kom=2,  nama="Tepung terigu",      unit="kg",
         kualitas_col="AB", harga_col="AC", tanggal_col="AD", asal_col="AE",
         choice_map={1: 4, 2: 5, 3: 6, 4: 19},
         choice_names={1: "Cakra Kembar", 2: "Segitiga Kemasan", 3: "Segitiga Biru Curah", 4: "Lainnya"}),
    dict(id_kom=3,  nama="Minyak Goreng",      unit="kg",
         kualitas_col="AF", harga_col="AG", tanggal_col="AH", asal_col="AI",
         choice_map={1: 7, 2: 8, 3: 9, 4: 20},
         choice_names={1: "Sunco", 2: "Fortune", 3: "Bimoli", 4: "Lainnya"}),
    dict(id_kom=4,  nama="Gula Pasir",         unit="kg",
         kualitas_col="AJ", harga_col="AK", tanggal_col="AL", asal_col="AM",  # <- fixed (was AI)
         choice_map={1: 10, 2: 12, 3: 21},
         choice_names={1: "Curah", 2: "PSM", 3: "Lainnya"}),
    dict(id_kom=5,  nama="Mie Kering Instan",  unit="bungkus",
         kualitas_col="AN", harga_col="AO", tanggal_col="AP", asal_col="AQ",
         choice_map={1: 13, 2: 23, 3: 24},
         choice_names={1: "Indomie Goreng", 2: "Indomie Kari Ayam", 3: "Mie sedap Goreng"}),
    dict(id_kom=6,  nama="Susu Bubuk Balita",  unit="Kotak 400 Gram",
         kualitas_col="AR", harga_col="AS", tanggal_col="AT", asal_col="AU",
         choice_map={1: 25, 2: 26, 3: 27},
         choice_names={1: "Frisian Flag 123", 2: "Dancow 1+", 3: "Lainnya"}),
    dict(id_kom=7,  nama="Susu Bubuk",         unit="Per-Kotak 400 Gram",
         kualitas_col="AV", harga_col="AW", tanggal_col="AX", asal_col="AY",
         choice_map={1: 28, 2: 29, 3: 30},
         choice_names={1: "Frisian Flag Full Cream", 2: "Dancow Instant", 3: "Lainnya"}),
    dict(id_kom=8,  nama="Cabai Merah",        unit="kg",
         kualitas_col="AZ", harga_col="BA", tanggal_col="BB", asal_col="BC",
         choice_map={1: 31, 2: 32},
         choice_names={1: "Keriting Segar", 2: "Besar Segar"}),
    dict(id_kom=9,  nama="Cabai Rawit",        unit="kg",
         kualitas_col=None, harga_col="BI", tanggal_col="BJ", asal_col="BK",
         fixed_h=33, fixed_i="Segar"),
    dict(id_kom=10, nama="Daging Ayam Ras",    unit="kg",
         kualitas_col=None, harga_col="BE", tanggal_col="BF", asal_col="BG",
         fixed_h=34, fixed_i="Segar"),
    dict(id_kom=11, nama="Telur Ayam Ras",     unit="kg",
         kualitas_col=None, harga_col="BM", tanggal_col="BN", asal_col="BO",
         fixed_h=35, fixed_i="Sedang"),
    dict(id_kom=12, nama="Bawang Merah",       unit="kg",
         kualitas_col=None, harga_col="BQ", tanggal_col="BR", asal_col="BS",
         fixed_h=36, fixed_i="Sedang"),
    dict(id_kom=13, nama="Bawang Putih",       unit="kg",
         kualitas_col=None, harga_col="BU", tanggal_col="BV", asal_col="BW",
         fixed_h=37, fixed_i="Sedang"),
    dict(id_kom=14, nama="Daging Sapi",        unit="kg",
         kualitas_col=None, harga_col="BY", tanggal_col="BZ", asal_col="CA",
         fixed_h=38, fixed_i="Segar"),
    dict(id_kom=15, nama="Udang",              unit="kg",
         kualitas_col=None, harga_col="CC", tanggal_col="CD", asal_col="CE",
         fixed_h=39, fixed_i="Sedang Segar"),
    dict(id_kom=16, nama="Ikan Kembung",       unit="kg",
         kualitas_col=None, harga_col="CG", tanggal_col="CH", asal_col="CI",
         fixed_h=40, fixed_i="Sedang Segar"),
    dict(id_kom=17, nama="Tempe",              unit="kg",
         kualitas_col=None, harga_col="CK", tanggal_col="CL", asal_col="CM",
         fixed_h=41, fixed_i="Putih Bersih"),
    dict(id_kom=18, nama="Tahu Mentah",        unit="kg",
         kualitas_col=None, harga_col="CO", tanggal_col="CP", asal_col="CQ",
         fixed_h=42, fixed_i="Putih Bersih"),
    dict(id_kom=19, nama="Pisang",             unit="kg",
         kualitas_col=None, harga_col="CS", tanggal_col="CT", asal_col="CU",
         fixed_h=43, fixed_i="Segar"),
    dict(id_kom=20, nama="Jeruk",              unit="kg",
         kualitas_col=None, harga_col="CW", tanggal_col="CX", asal_col="CY",
         fixed_h=44, fixed_i="Sedang Segar"),
]

BULAN_ID = ["Januari", "Februari", "Maret", "April", "Mei", "Juni", "Juli",
            "Agustus", "September", "Oktober", "November", "Desember"]
BULAN_ID_3 = [b[:3] for b in BULAN_ID]  # Jan, Feb, Mar, Apr, Mei, Jun, Jul, Agu, Sep, Okt, Nov, Des

CHOICE_RE = re.compile(r"^\s*(\d+)\s*\.")


def excel_weeknum(d: datetime.date) -> int:
    """Replikasi Excel WEEKNUM(date,1): minggu mulai hari Minggu, minggu 1 selalu berisi 1 Jan."""
    jan1 = datetime.date(d.year, 1, 1)
    py_wd_jan1 = jan1.weekday()  # Senin=0 ... Minggu=6
    excel_wd_jan1 = ((py_wd_jan1 + 1) % 7) + 1  # Minggu=1 ... Sabtu=7
    days_since_jan1 = (d - jan1).days
    return (days_since_jan1 + excel_wd_jan1 - 1) // 7 + 1


def week_of_month(d: datetime.date) -> int:
    first_of_month = d.replace(day=1)
    return excel_weeknum(d) - excel_weeknum(first_of_month) + 1


def parse_choice(raw_text):
    if raw_text is None:
        return None
    s = str(raw_text).strip()
    if not s:
        return None
    m = CHOICE_RE.match(s)
    return int(m.group(1)) if m else None


def load_quality_names(wb):
    """id_kualitas -> nama_kualitas, dari sheet 'Kode Kualitas'."""
    ws = wb[QUALITY_SHEET]
    mapping = {}
    header = [c.value for c in ws[1]]
    idx_id = header.index("id_kualitas")
    idx_nama = header.index("nama_kualitas")
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx_id] is None:
            continue
        mapping[int(row[idx_id])] = row[idx_nama]
    return mapping


def generate(input_path: str, output_path: str):
    wb = openpyxl.load_workbook(input_path, data_only=True)
    ws3 = wb[SRC_SHEET]
    quality_names = load_quality_names(wb)

    # deteksi baris terakhir data responden (berdasarkan kolom A / NIK)
    last_row = SRC_DATA_START_ROW - 1
    for r in range(SRC_DATA_START_ROW, ws3.max_row + 1):
        if ws3.cell(row=r, column=1).value not in (None, ""):
            last_row = r
    n_rows = last_row - SRC_DATA_START_ROW + 1
    print(f"Data responden terdeteksi: {n_rows} baris (sheet3 baris {SRC_DATA_START_ROW}-{last_row})")

    # buat ulang sheet dashboard
    if DASHBOARD_SHEET in wb.sheetnames:
        del wb[DASHBOARD_SHEET]
    ws4 = wb.create_sheet(DASHBOARD_SHEET)

    headers = ["Id Kec Desa", "Kode Kecamatan", "Kecamatan", "Kode Desa", "Desa",
               "Id Komoditas", "Komoditas", "Id Kualitas", "Kualitas", "Satuan",
               "Harga", "Tanggal Pendataan", "Minggu", "No Bulan", "Bulan",
               "Periode", "No Periode", "Asal Barang",
               "Periode (Tahun-Bulan-Minggu)", "Tahun"]
    ws4.append(headers)

    total_written = 0
    for block in BLOCKS:
        id_kom = block["id_kom"]
        nama = block["nama"]
        unit = block["unit"]
        harga_col = column_index_from_string(block["harga_col"])
        tanggal_col = column_index_from_string(block["tanggal_col"])
        asal_col = column_index_from_string(block["asal_col"])
        kualitas_col = column_index_from_string(block["kualitas_col"]) if block["kualitas_col"] else None

        for r in range(SRC_DATA_START_ROW, last_row + 1):
            id_vals = [ws3.cell(row=r, column=column_index_from_string(c)).value for c in ID_COLS]

            harga = ws3.cell(row=r, column=harga_col).value
            tanggal_raw = ws3.cell(row=r, column=tanggal_col).value
            asal = ws3.cell(row=r, column=asal_col).value

            # tentukan id_kualitas & nama_kualitas
            if kualitas_col is not None:
                raw_q = ws3.cell(row=r, column=kualitas_col).value
                choice = parse_choice(raw_q)
                id_kualitas = block["choice_map"].get(choice) if choice else None
                nama_kualitas = block["choice_names"].get(choice) if choice else None
            else:
                id_kualitas = block["fixed_h"]
                nama_kualitas = block["fixed_i"]

            # kalau tidak ada tanggal -> tidak ada data untuk komoditas ini di responden itu,
            # kosongkan kolom turunan tanggal (beda dgn template asli yg malah error #VALUE!)
            tanggal = None
            minggu = no_bulan = bulan_nama = periode = no_periode = periode_full = tahun = None
            d = None
            if isinstance(tanggal_raw, (datetime.date, datetime.datetime)):
                d = tanggal_raw.date() if isinstance(tanggal_raw, datetime.datetime) else tanggal_raw
            elif isinstance(tanggal_raw, str) and tanggal_raw.strip():
                try:
                    d = datetime.datetime.fromisoformat(tanggal_raw.strip()).date()
                except ValueError:
                    d = None
            if d is not None:
                tanggal = d
                minggu = week_of_month(d)
                no_bulan = f"{d.month:02d}"
                bulan_nama = BULAN_ID[d.month - 1]
                periode = f"{BULAN_ID_3[d.month - 1]}-{minggu}"
                no_periode = f"{no_bulan}-{minggu}"
                tahun = d.year
                periode_full = f"{tahun}-{no_periode}"

            row_out = [
                *id_vals,
                id_kom, nama,
                id_kualitas, nama_kualitas,
                unit,
                harga,
                tanggal,
                minggu, no_bulan, bulan_nama, periode, no_periode,
                asal,
                periode_full, tahun,
            ]
            ws4.append(row_out)
            total_written += 1

    # format kolom tanggal
    for row in ws4.iter_rows(min_row=2, min_col=12, max_col=12):
        for cell in row:
            if cell.value is not None:
                cell.number_format = "yyyy-mm-dd"

    print(f"Total baris ditulis ke '{DASHBOARD_SHEET}': {total_written} (20 komoditas x {n_rows} responden)")
    wb.save(output_path)
    print(f"Selesai. File disimpan: {output_path}")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python3 generate_dashboard.py <input.xlsx> <output.xlsx>")
        sys.exit(1)
    generate(sys.argv[1], sys.argv[2])