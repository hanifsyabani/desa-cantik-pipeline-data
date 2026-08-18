"""
generate_full_pipeline.py
--------------------------
Otomasi PENUH dari sheet "1. list_data (ori)" (data mentah hasil download
aplikasi kuisioner) langsung ke hasil akhir siap upload -- menggantikan
SELURUH proses manual sheet 2 (bersihin kode wilayah) -> sheet 3 (paste
value) -> sheet 4 (unpivot 20 komoditas) -> sheet 5 (paste value) ->
sheet 6 (filter data valid).

CARA PAKAI:
    python3 generate_full_pipeline.py <file_input.xlsx> <file_output.xlsx> [--min-harga N]

Contoh:
    python3 generate_full_pipeline.py data_minggu_ini.xlsx hasil.xlsx
    python3 generate_full_pipeline.py data_minggu_ini.xlsx hasil.xlsx --min-harga 500

Output berisi 2 sheet:
    - "Dashboard (Raw)"     -> setara sheet 5, semua baris (termasuk yang
                               kosong/gagal lookup wilayah, ditandai kosong
                               bukan #VALUE! seperti file lama)
    - "Dashboard (Cleaned)" -> setara sheet 6, sudah difilter:
                               * buang baris tanpa jawaban (kualitas/harga/tanggal kosong)
                               * buang baris yang kode wilayahnya gagal di-lookup (#N/A)
                               * buang harga di bawah --min-harga (default 1000,
                                 mengantisipasi salah ketik seperti "13" yang
                                 harusnya "13000")

File input CUKUP berisi sheet "1. list_data (ori)" + 2 sheet referensi yang
tidak berubah: "kd wilayah" dan "Kode Kualitas".
"""

import sys
import re
import argparse
import datetime
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

SRC_SHEET = "1. list_data (ori)"
WILAYAH_SHEET = "kd wilayah"
QUALITY_SHEET = "Kode Kualitas"

SRC_HEADER_ROW = 7
SRC_DATA_START_ROW = 8

# Kolom mentah di sheet 1 (raw)
COL_KECAMATAN_RAW = 9   # I: "KECAMATAN SEKAYU"
COL_DESA_RAW = 10       # J: "DESA SUNGAI MEDAK"

# Konfigurasi 20 blok komoditas -- kolom dirujuk relatif terhadap sheet
# "3.list_data (value)" (huruf di bawah), lalu otomatis digeser -7 kolom
# untuk mendapat posisi aslinya di sheet 1. Lihat generate_dashboard.py
# untuk penjelasan choice_map / fixed_h / fixed_i.
_COL_SHIFT_SHEET3_TO_SHEET1 = -7

BLOCKS = [
    dict(id_kom=1,  nama="Beras",              unit="kg",
         kualitas_col="X",  harga_col="Y",  tanggal_col="Z",  asal_col="AA",
         choice_map={1: 1, 2: 3, 3: 18},
         choice_names={1: "Selancar", 2: "Topi Koki", 3: "Lainnya"}),
    dict(id_kom=2,  nama="Tepung terigu",      unit="kg",
         kualitas_col="AB", harga_col="AC", tanggal_col="AD", asal_col="AE",
         choice_map={1: 4, 2: 5, 3: 6, 4: 19},
         choice_names={1: "Cakra Kembar", 2: "Segitiga Kemasan", 3: "Segitiga Biru Curah", 4: "Lainnya"}),
    dict(id_kom=3,  nama="Minyak Goreng",      unit="kg",
         kualitas_col="AF", harga_col="AG", tanggal_col="AH", asal_col="AI",
         choice_map={1: 7, 2: 8, 3: 9, 4: 20},
         choice_names={1: "Sunco", 2: "Fortune", 3: "Bimoli", 4: "Lainnya"}),
    dict(id_kom=4,  nama="Gula Pasir",         unit="kg",
         kualitas_col="AJ", harga_col="AK", tanggal_col="AL", asal_col="AM",  # bug lama (AI) sudah diperbaiki
         choice_map={1: 10, 2: 12, 3: 21},
         choice_names={1: "Curah", 2: "PSM", 3: "Lainnya"}),
    dict(id_kom=5,  nama="Mie Kering Instan",  unit="bungkus",
         kualitas_col="AN", harga_col="AO", tanggal_col="AP", asal_col="AQ",
         choice_map={1: 13, 2: 23, 3: 24},
         choice_names={1: "Indomie Goreng", 2: "Indomie Kari Ayam", 3: "Mie sedap Goreng"}),
    dict(id_kom=6,  nama="Susu Bubuk Balita",  unit="Kotak 400 Gram",
         kualitas_col="AR", harga_col="AS", tanggal_col="AT", asal_col="AU",
         choice_map={1: 25, 2: 26, 3: 27},
         choice_names={1: "Frisian Flag 123", 2: "Dancow 1+", 3: "Lainnya"}),
    dict(id_kom=7,  nama="Susu Bubuk",         unit="Per-Kotak 400 Gram",
         kualitas_col="AV", harga_col="AW", tanggal_col="AX", asal_col="AY",
         choice_map={1: 28, 2: 29, 3: 30},
         choice_names={1: "Frisian Flag Full Cream", 2: "Dancow Instant", 3: "Lainnya"}),
    dict(id_kom=8,  nama="Cabai Merah",        unit="kg",
         kualitas_col="AZ", harga_col="BA", tanggal_col="BB", asal_col="BC",
         choice_map={1: 31, 2: 32},
         choice_names={1: "Keriting Segar", 2: "Besar Segar"}),
    dict(id_kom=9,  nama="Cabai Rawit",        unit="kg",
         kualitas_col=None, harga_col="BI", tanggal_col="BJ", asal_col="BK",
         fixed_h=33, fixed_i="Segar"),
    dict(id_kom=10, nama="Daging Ayam Ras",    unit="kg",
         kualitas_col=None, harga_col="BE", tanggal_col="BF", asal_col="BG",
         fixed_h=34, fixed_i="Segar"),
    dict(id_kom=11, nama="Telur Ayam Ras",     unit="kg",
         kualitas_col=None, harga_col="BM", tanggal_col="BN", asal_col="BO",
         fixed_h=35, fixed_i="Sedang"),
    dict(id_kom=12, nama="Bawang Merah",       unit="kg",
         kualitas_col=None, harga_col="BQ", tanggal_col="BR", asal_col="BS",
         fixed_h=36, fixed_i="Sedang"),
    dict(id_kom=13, nama="Bawang Putih",       unit="kg",
         kualitas_col=None, harga_col="BU", tanggal_col="BV", asal_col="BW",
         fixed_h=37, fixed_i="Sedang"),
    dict(id_kom=14, nama="Daging Sapi",        unit="kg",
         kualitas_col=None, harga_col="BY", tanggal_col="BZ", asal_col="CA",
         fixed_h=38, fixed_i="Segar"),
    dict(id_kom=15, nama="Udang",              unit="kg",
         kualitas_col=None, harga_col="CC", tanggal_col="CD", asal_col="CE",
         fixed_h=39, fixed_i="Sedang Segar"),
    dict(id_kom=16, nama="Ikan Kembung",       unit="kg",
         kualitas_col=None, harga_col="CG", tanggal_col="CH", asal_col="CI",
         fixed_h=40, fixed_i="Sedang Segar"),
    dict(id_kom=17, nama="Tempe",              unit="kg",
         kualitas_col=None, harga_col="CK", tanggal_col="CL", asal_col="CM",
         fixed_h=41, fixed_i="Putih Bersih"),
    dict(id_kom=18, nama="Tahu Mentah",        unit="kg",
         kualitas_col=None, harga_col="CO", tanggal_col="CP", asal_col="CQ",
         fixed_h=42, fixed_i="Putih Bersih"),
    dict(id_kom=19, nama="Pisang",             unit="kg",
         kualitas_col=None, harga_col="CS", tanggal_col="CT", asal_col="CU",
         fixed_h=43, fixed_i="Segar"),
    dict(id_kom=20, nama="Jeruk",              unit="kg",
         kualitas_col=None, harga_col="CW", tanggal_col="CX", asal_col="CY",
         fixed_h=44, fixed_i="Sedang Segar"),
]


def shifted(col_letter):
    """Geser huruf kolom (posisi di sheet3) -7 supaya jadi posisi di sheet1 (ori)."""
    idx = column_index_from_string(col_letter) + _COL_SHIFT_SHEET3_TO_SHEET1
    return idx


BULAN_ID = ["Januari", "Februari", "Maret", "April", "Mei", "Juni", "Juli",
            "Agustus", "September", "Oktober", "November", "Desember"]
BULAN_ID_3 = [b[:3] for b in BULAN_ID]

CHOICE_RE = re.compile(r"^\s*(\d+)\s*\.")


def excel_weeknum(d: datetime.date) -> int:
    jan1 = datetime.date(d.year, 1, 1)
    py_wd_jan1 = jan1.weekday()
    excel_wd_jan1 = ((py_wd_jan1 + 1) % 7) + 1
    days_since_jan1 = (d - jan1).days
    return (days_since_jan1 + excel_wd_jan1 - 1) // 7 + 1


def week_of_month(d: datetime.date) -> int:
    first_of_month = d.replace(day=1)
    return excel_weeknum(d) - excel_weeknum(first_of_month) + 1


def parse_choice(raw_text):
    if raw_text is None:
        return None
    s = str(raw_text).strip()
    if not s:
        return None
    m = CHOICE_RE.match(s)
    return int(m.group(1)) if m else None


def parse_date(raw):
    if isinstance(raw, (datetime.date, datetime.datetime)):
        return raw.date() if isinstance(raw, datetime.datetime) else raw
    if isinstance(raw, str) and raw.strip():
        try:
            return datetime.datetime.fromisoformat(raw.strip()).date()
        except ValueError:
            return None
    return None


def load_wilayah_lookup(wb):
    """Bangun dict lookup dari sheet 'kd wilayah', meniru rumus SUBSTITUTE/VLOOKUP di sheet2."""
    ws = wb[WILAYAH_SHEET]
    kec_name_to_code = {}   # KECAMATAN (dedup, kol O) -> Kode Kec BPS (kol P)
    kec_code_to_name = {}   # Kode Kec BPS (kol N) -> KECAMATAN (kol O)
    desa_nospace_to_code = {}  # DESA TANPA SPASI (kol I) -> Kode Desa BPS (kol J)
    idbps_to_desaname = {}  # Id BPS (kol D) -> DESA (kol H)

    for row in ws.iter_rows(min_row=2, values_only=True):
        # kolom index (0-based dalam tuple row): D=3,H=7,I=8,J=9,N=13,O=14,P=15
        id_bps, desa_h = row[3], row[7]
        if id_bps is not None and desa_h is not None:
            idbps_to_desaname[str(id_bps)] = desa_h

        desa_nospace, kode_desa_bps = row[8], row[9]
        if desa_nospace is not None and kode_desa_bps is not None:
            desa_nospace_to_code[str(desa_nospace)] = kode_desa_bps

        kode_kec_bps_n, kec_o = row[13], row[14]
        if kode_kec_bps_n is not None and kec_o is not None:
            kec_code_to_name[str(kode_kec_bps_n)] = kec_o

        kec_o2, kode_kec_bps_p = row[14], row[15]
        if kec_o2 is not None and kode_kec_bps_p is not None:
            kec_name_to_code[str(kec_o2)] = kode_kec_bps_p

    return kec_name_to_code, kec_code_to_name, desa_nospace_to_code, idbps_to_desaname


def load_quality_names(wb):
    ws = wb[QUALITY_SHEET]
    mapping = {}
    header = [c.value for c in ws[1]]
    idx_id = header.index("id_kualitas")
    idx_nama = header.index("nama_kualitas")
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx_id] is None:
            continue
        mapping[int(row[idx_id])] = row[idx_nama]
    return mapping


def clean_wilayah(kec_raw, desa_raw, lookups):
    kec_name_to_code, kec_code_to_name, desa_nospace_to_code, idbps_to_desaname = lookups

    kec_clean = (kec_raw or "").replace("KECAMATAN ", "").strip()
    desa_step1 = (desa_raw or "").replace("DESA", "")
    desa_step2 = desa_step1.replace("KELURAHAN", "")
    desa_nospace = desa_step2.replace(" ", "")

    kec_code = kec_name_to_code.get(kec_clean)
    kec_final = kec_code_to_name.get(str(kec_code)) if kec_code is not None else None
    desa_code = desa_nospace_to_code.get(desa_nospace)

    if kec_code is not None and desa_code is not None:
        id_kec_desa = f"{kec_code}{desa_code}"
    else:
        id_kec_desa = "#N/A"

    desa_final = idbps_to_desaname.get(id_kec_desa)

    valid = kec_code is not None and desa_code is not None and desa_final is not None
    return dict(
        id_kec_desa=id_kec_desa if valid else "#N/A",
        kode_kecamatan=kec_code if kec_code is not None else "#N/A",
        kecamatan=kec_final if kec_final else "#N/A",
        kode_desa=desa_code if desa_code is not None else "#N/A",
        desa=desa_final if desa_final else "#N/A",
        valid=valid,
    )


def generate(input_path: str, output_path: str, min_harga: int):
    wb = openpyxl.load_workbook(input_path, data_only=True)
    ws1 = wb[SRC_SHEET]
    lookups = load_wilayah_lookup(wb)

    last_row = SRC_DATA_START_ROW - 1
    for r in range(SRC_DATA_START_ROW, ws1.max_row + 1):
        if ws1.cell(row=r, column=1).value not in (None, ""):
            last_row = r
    n_rows = last_row - SRC_DATA_START_ROW + 1
    print(f"Data responden terdeteksi: {n_rows} baris (sheet1 baris {SRC_DATA_START_ROW}-{last_row})")

    out_wb = openpyxl.Workbook()
    ws_raw = out_wb.active
    ws_raw.title = "Dashboard (Raw)"
    ws_clean = out_wb.create_sheet("Dashboard (Cleaned)")

    headers = ["Id Kec Desa", "Kode Kecamatan", "Kecamatan", "Kode Desa", "Desa",
               "Id Komoditas", "Komoditas", "Id Kualitas", "Kualitas", "Satuan",
               "Harga", "Tanggal Pendataan", "Minggu", "No Bulan", "Bulan",
               "Periode", "No Periode", "Asal Barang",
               "Periode (Tahun-Bulan-Minggu)", "Tahun"]
    ws_raw.append(headers)
    ws_clean.append(headers)

    # cache hasil cleaning wilayah per baris responden (dipakai berulang utk 20 komoditas)
    wilayah_cache = {}
    for r in range(SRC_DATA_START_ROW, last_row + 1):
        kec_raw = ws1.cell(row=r, column=COL_KECAMATAN_RAW).value
        desa_raw = ws1.cell(row=r, column=COL_DESA_RAW).value
        wilayah_cache[r] = clean_wilayah(kec_raw, desa_raw, lookups)

    n_raw = n_clean = n_dropped_wilayah = n_dropped_price = n_blank = 0

    for block in BLOCKS:
        id_kom = block["id_kom"]
        nama = block["nama"]
        unit = block["unit"]
        harga_col = shifted(block["harga_col"])
        tanggal_col = shifted(block["tanggal_col"])
        asal_col = shifted(block["asal_col"])
        kualitas_col = shifted(block["kualitas_col"]) if block["kualitas_col"] else None

        for r in range(SRC_DATA_START_ROW, last_row + 1):
            wv = wilayah_cache[r]

            harga = ws1.cell(row=r, column=harga_col).value
            tanggal_raw = ws1.cell(row=r, column=tanggal_col).value
            asal = ws1.cell(row=r, column=asal_col).value

            if kualitas_col is not None:
                raw_q = ws1.cell(row=r, column=kualitas_col).value
                choice = parse_choice(raw_q)
                id_kualitas = block["choice_map"].get(choice) if choice else None
                nama_kualitas = block["choice_names"].get(choice) if choice else None
            else:
                id_kualitas = block["fixed_h"]
                nama_kualitas = block["fixed_i"]

            d = parse_date(tanggal_raw)
            tanggal = minggu = no_bulan = bulan_nama = periode = no_periode = periode_full = tahun = None
            if d is not None:
                tanggal = d
                minggu = week_of_month(d)
                no_bulan = f"{d.month:02d}"
                bulan_nama = BULAN_ID[d.month - 1]
                periode = f"{BULAN_ID_3[d.month - 1]}-{minggu}"
                no_periode = f"{no_bulan}-{minggu}"
                tahun = d.year
                periode_full = f"{tahun}-{no_periode}"

            row_out = [
                wv["id_kec_desa"], wv["kode_kecamatan"], wv["kecamatan"], wv["kode_desa"], wv["desa"],
                id_kom, nama,
                id_kualitas, nama_kualitas,
                unit,
                harga,
                tanggal,
                minggu, no_bulan, bulan_nama, periode, no_periode,
                asal,
                periode_full, tahun,
            ]
            ws_raw.append(row_out)
            n_raw += 1

            # ---- filter untuk versi "Cleaned" ----
            is_blank = (id_kualitas is None and harga in (None, 0, "") and d is None)
            if is_blank:
                n_blank += 1
                continue
            if not wv["valid"]:
                n_dropped_wilayah += 1
                continue
            if harga is None or (isinstance(harga, (int, float)) and harga < min_harga):
                n_dropped_price += 1
                continue
            ws_clean.append(row_out)
            n_clean += 1

    for ws in (ws_raw, ws_clean):
        for row in ws.iter_rows(min_row=2, min_col=12, max_col=12):
            for cell in row:
                if cell.value is not None:
                    cell.number_format = "yyyy-mm-dd"

    print(f"'Dashboard (Raw)': {n_raw} baris")
    print(f"'Dashboard (Cleaned)': {n_clean} baris")
    print(f"  - dibuang (tanpa jawaban): {n_blank}")
    print(f"  - dibuang (wilayah gagal lookup / #N/A): {n_dropped_wilayah}")
    print(f"  - dibuang (harga < {min_harga}, dianggap salah ketik): {n_dropped_price}")

    out_wb.save(output_path)
    print(f"Selesai. File disimpan: {output_path}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("input")
    parser.add_argument("output")
    parser.add_argument("--min-harga", type=int, default=1000,
                         help="Harga di bawah ini dibuang dari sheet Cleaned (default 1000)")
    args = parser.parse_args()
    generate(args.input, args.output, args.min_harga)
