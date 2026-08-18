"""
streamlit_app.py
------------------
Versi web (Streamlit) dari generate_full_pipeline.py.

Otomasi PENUH dari sheet "1. list_data (ori)" (data mentah hasil download
aplikasi kuisioner) langsung ke hasil akhir siap upload -- menggantikan
SELURUH proses manual sheet 2 (bersihin kode wilayah) -> sheet 3 (paste
value) -> sheet 4 (unpivot 20 komoditas) -> sheet 5 (paste value) ->
sheet 6 (filter data valid).

CARA JALANIN:
    pip install streamlit openpyxl
    streamlit run streamlit_app.py

Lalu browser otomatis kebuka. Tinggal upload file .xlsx yang berisi sheet
"1. list_data (ori)" + 2 sheet referensi "kd wilayah" dan "Kode Kualitas",
atur minimum harga kalau perlu, klik "Proses", lalu download hasilnya.
"""

import re
import io
import datetime

import streamlit as st
import openpyxl
from openpyxl.utils import column_index_from_string

# ============================================================
# ============ LOGIC INTI (sama persis dgn versi CLI) ========
# ============================================================

SRC_SHEET = "1. list_data (ori)"
WILAYAH_SHEET = "kd wilayah"
QUALITY_SHEET = "Kode Kualitas"

SRC_HEADER_ROW = 7
SRC_DATA_START_ROW = 8

COL_KECAMATAN_RAW = 9   # I: "KECAMATAN SEKAYU"
COL_DESA_RAW = 10       # J: "DESA SUNGAI MEDAK"

_COL_SHIFT_SHEET3_TO_SHEET1 = -7

BLOCKS = [
    dict(id_kom=1,  nama="Beras",              unit="kg",
         kualitas_col="X",  harga_col="Y",  tanggal_col="Z",  asal_col="AA",
         choice_map={1: 1, 2: 3, 3: 18},
         choice_names={1: "Selancar", 2: "Topi Koki", 3: "Lainnya"}),
    dict(id_kom=2,  nama="Tepung terigu",      unit="kg",
         kualitas_col="AB", harga_col="AC", tanggal_col="AD", asal_col="AE",
         choice_map={1: 4, 2: 5, 3: 6, 4: 19},
         choice_names={1: "Cakra Kembar", 2: "Segitiga Kemasan", 3: "Segitiga Biru Curah", 4: "Lainnya"}),
    dict(id_kom=3,  nama="Minyak Goreng",      unit="kg",
         kualitas_col="AF", harga_col="AG", tanggal_col="AH", asal_col="AI",
         choice_map={1: 7, 2: 8, 3: 9, 4: 20},
         choice_names={1: "Sunco", 2: "Fortune", 3: "Bimoli", 4: "Lainnya"}),
    dict(id_kom=4,  nama="Gula Pasir",         unit="kg",
         kualitas_col="AJ", harga_col="AK", tanggal_col="AL", asal_col="AM",
         choice_map={1: 10, 2: 12, 3: 21},
         choice_names={1: "Curah", 2: "PSM", 3: "Lainnya"}),
    dict(id_kom=5,  nama="Mie Kering Instan",  unit="bungkus",
         kualitas_col="AN", harga_col="AO", tanggal_col="AP", asal_col="AQ",
         choice_map={1: 13, 2: 23, 3: 24},
         choice_names={1: "Indomie Goreng", 2: "Indomie Kari Ayam", 3: "Mie sedap Goreng"}),
    dict(id_kom=6,  nama="Susu Bubuk Balita",  unit="Kotak 400 Gram",
         kualitas_col="AR", harga_col="AS", tanggal_col="AT", asal_col="AU",
         choice_map={1: 25, 2: 26, 3: 27},
         choice_names={1: "Frisian Flag 123", 2: "Dancow 1+", 3: "Lainnya"}),
    dict(id_kom=7,  nama="Susu Bubuk",         unit="Per-Kotak 400 Gram",
         kualitas_col="AV", harga_col="AW", tanggal_col="AX", asal_col="AY",
         choice_map={1: 28, 2: 29, 3: 30},
         choice_names={1: "Frisian Flag Full Cream", 2: "Dancow Instant", 3: "Lainnya"}),
    dict(id_kom=8,  nama="Cabai Merah",        unit="kg",
         kualitas_col="AZ", harga_col="BA", tanggal_col="BB", asal_col="BC",
         choice_map={1: 31, 2: 32},
         choice_names={1: "Keriting Segar", 2: "Besar Segar"}),
    dict(id_kom=9,  nama="Cabai Rawit",        unit="kg",
         kualitas_col=None, harga_col="BI", tanggal_col="BJ", asal_col="BK",
         fixed_h=33, fixed_i="Segar"),
    dict(id_kom=10, nama="Daging Ayam Ras",    unit="kg",
         kualitas_col=None, harga_col="BE", tanggal_col="BF", asal_col="BG",
         fixed_h=34, fixed_i="Segar"),
    dict(id_kom=11, nama="Telur Ayam Ras",     unit="kg",
         kualitas_col=None, harga_col="BM", tanggal_col="BN", asal_col="BO",
         fixed_h=35, fixed_i="Sedang"),
    dict(id_kom=12, nama="Bawang Merah",       unit="kg",
         kualitas_col=None, harga_col="BQ", tanggal_col="BR", asal_col="BS",
         fixed_h=36, fixed_i="Sedang"),
    dict(id_kom=13, nama="Bawang Putih",       unit="kg",
         kualitas_col=None, harga_col="BU", tanggal_col="BV", asal_col="BW",
         fixed_h=37, fixed_i="Sedang"),
    dict(id_kom=14, nama="Daging Sapi",        unit="kg",
         kualitas_col=None, harga_col="BY", tanggal_col="BZ", asal_col="CA",
         fixed_h=38, fixed_i="Segar"),
    dict(id_kom=15, nama="Udang",              unit="kg",
         kualitas_col=None, harga_col="CC", tanggal_col="CD", asal_col="CE",
         fixed_h=39, fixed_i="Sedang Segar"),
    dict(id_kom=16, nama="Ikan Kembung",       unit="kg",
         kualitas_col=None, harga_col="CG", tanggal_col="CH", asal_col="CI",
         fixed_h=40, fixed_i="Sedang Segar"),
    dict(id_kom=17, nama="Tempe",              unit="kg",
         kualitas_col=None, harga_col="CK", tanggal_col="CL", asal_col="CM",
         fixed_h=41, fixed_i="Putih Bersih"),
    dict(id_kom=18, nama="Tahu Mentah",        unit="kg",
         kualitas_col=None, harga_col="CO", tanggal_col="CP", asal_col="CQ",
         fixed_h=42, fixed_i="Putih Bersih"),
    dict(id_kom=19, nama="Pisang",             unit="kg",
         kualitas_col=None, harga_col="CS", tanggal_col="CT", asal_col="CU",
         fixed_h=43, fixed_i="Segar"),
    dict(id_kom=20, nama="Jeruk",              unit="kg",
         kualitas_col=None, harga_col="CW", tanggal_col="CX", asal_col="CY",
         fixed_h=44, fixed_i="Sedang Segar"),
]

BULAN_ID = ["Januari", "Februari", "Maret", "April", "Mei", "Juni", "Juli",
            "Agustus", "September", "Oktober", "November", "Desember"]
BULAN_ID_3 = [b[:3] for b in BULAN_ID]

CHOICE_RE = re.compile(r"^\s*(\d+)\s*\.")


def shifted(col_letter):
    return column_index_from_string(col_letter) + _COL_SHIFT_SHEET3_TO_SHEET1


def excel_weeknum(d: datetime.date) -> int:
    jan1 = datetime.date(d.year, 1, 1)
    py_wd_jan1 = jan1.weekday()
    excel_wd_jan1 = ((py_wd_jan1 + 1) % 7) + 1
    days_since_jan1 = (d - jan1).days
    return (days_since_jan1 + excel_wd_jan1 - 1) // 7 + 1


def week_of_month(d: datetime.date) -> int:
    first_of_month = d.replace(day=1)
    return excel_weeknum(d) - excel_weeknum(first_of_month) + 1


def parse_choice(raw_text):
    if raw_text is None:
        return None
    s = str(raw_text).strip()
    if not s:
        return None
    m = CHOICE_RE.match(s)
    return int(m.group(1)) if m else None


def parse_date(raw):
    if isinstance(raw, (datetime.date, datetime.datetime)):
        return raw.date() if isinstance(raw, datetime.datetime) else raw
    if isinstance(raw, str) and raw.strip():
        try:
            return datetime.datetime.fromisoformat(raw.strip()).date()
        except ValueError:
            return None
    return None


def load_wilayah_lookup(wb):
    ws = wb[WILAYAH_SHEET]
    kec_name_to_code = {}
    kec_code_to_name = {}
    desa_nospace_to_code = {}
    idbps_to_desaname = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        id_bps, desa_h = row[3], row[7]
        if id_bps is not None and desa_h is not None:
            idbps_to_desaname[str(id_bps)] = desa_h

        desa_nospace, kode_desa_bps = row[8], row[9]
        if desa_nospace is not None and kode_desa_bps is not None:
            desa_nospace_to_code[str(desa_nospace)] = kode_desa_bps

        kode_kec_bps_n, kec_o = row[13], row[14]
        if kode_kec_bps_n is not None and kec_o is not None:
            kec_code_to_name[str(kode_kec_bps_n)] = kec_o

        kec_o2, kode_kec_bps_p = row[14], row[15]
        if kec_o2 is not None and kode_kec_bps_p is not None:
            kec_name_to_code[str(kec_o2)] = kode_kec_bps_p

    return kec_name_to_code, kec_code_to_name, desa_nospace_to_code, idbps_to_desaname


def clean_wilayah(kec_raw, desa_raw, lookups):
    kec_name_to_code, kec_code_to_name, desa_nospace_to_code, idbps_to_desaname = lookups

    kec_clean = (kec_raw or "").replace("KECAMATAN ", "").strip()
    desa_step1 = (desa_raw or "").replace("DESA", "")
    desa_step2 = desa_step1.replace("KELURAHAN", "")
    desa_nospace = desa_step2.replace(" ", "")

    kec_code = kec_name_to_code.get(kec_clean)
    kec_final = kec_code_to_name.get(str(kec_code)) if kec_code is not None else None
    desa_code = desa_nospace_to_code.get(desa_nospace)

    if kec_code is not None and desa_code is not None:
        id_kec_desa = f"{kec_code}{desa_code}"
    else:
        id_kec_desa = "#N/A"

    desa_final = idbps_to_desaname.get(id_kec_desa)

    valid = kec_code is not None and desa_code is not None and desa_final is not None
    return dict(
        id_kec_desa=id_kec_desa if valid else "#N/A",
        kode_kecamatan=kec_code if kec_code is not None else "#N/A",
        kecamatan=kec_final if kec_final else "#N/A",
        kode_desa=desa_code if desa_code is not None else "#N/A",
        desa=desa_final if desa_final else "#N/A",
        valid=valid,
    )


def generate(input_bytes: bytes, min_harga: int, progress_cb=None):
    """Jalankan pipeline penuh, kembalikan (output_bytes, stats dict)."""
    wb = openpyxl.load_workbook(io.BytesIO(input_bytes), data_only=True)

    missing = [s for s in (SRC_SHEET, WILAYAH_SHEET, QUALITY_SHEET) if s not in wb.sheetnames]
    if missing:
        raise ValueError(
            "Sheet berikut tidak ditemukan di file: " + ", ".join(missing) +
            f"\nSheet yang ada: {wb.sheetnames}"
        )

    ws1 = wb[SRC_SHEET]
    lookups = load_wilayah_lookup(wb)

    last_row = SRC_DATA_START_ROW - 1
    for r in range(SRC_DATA_START_ROW, ws1.max_row + 1):
        if ws1.cell(row=r, column=1).value not in (None, ""):
            last_row = r
    n_rows = last_row - SRC_DATA_START_ROW + 1

    out_wb = openpyxl.Workbook()
    ws_raw = out_wb.active
    ws_raw.title = "Dashboard (Raw)"
    ws_clean = out_wb.create_sheet("Dashboard (Cleaned)")

    headers = ["Id Kec Desa", "Kode Kecamatan", "Kecamatan", "Kode Desa", "Desa",
               "Id Komoditas", "Komoditas", "Id Kualitas", "Kualitas", "Satuan",
               "Harga", "Tanggal Pendataan", "Minggu", "No Bulan", "Bulan",
               "Periode", "No Periode", "Asal Barang",
               "Periode (Tahun-Bulan-Minggu)", "Tahun"]
    ws_raw.append(headers)
    ws_clean.append(headers)

    wilayah_cache = {}
    for r in range(SRC_DATA_START_ROW, last_row + 1):
        kec_raw = ws1.cell(row=r, column=COL_KECAMATAN_RAW).value
        desa_raw = ws1.cell(row=r, column=COL_DESA_RAW).value
        wilayah_cache[r] = clean_wilayah(kec_raw, desa_raw, lookups)

    n_raw = n_clean = n_dropped_wilayah = n_dropped_price = n_blank = 0

    total_blocks = len(BLOCKS)
    for bi, block in enumerate(BLOCKS):
        if progress_cb:
            progress_cb((bi + 1) / total_blocks, block["nama"])

        id_kom = block["id_kom"]
        nama = block["nama"]
        unit = block["unit"]
        harga_col = shifted(block["harga_col"])
        tanggal_col = shifted(block["tanggal_col"])
        asal_col = shifted(block["asal_col"])
        kualitas_col = shifted(block["kualitas_col"]) if block["kualitas_col"] else None

        for r in range(SRC_DATA_START_ROW, last_row + 1):
            wv = wilayah_cache[r]

            harga = ws1.cell(row=r, column=harga_col).value
            tanggal_raw = ws1.cell(row=r, column=tanggal_col).value
            asal = ws1.cell(row=r, column=asal_col).value

            if kualitas_col is not None:
                raw_q = ws1.cell(row=r, column=kualitas_col).value
                choice = parse_choice(raw_q)
                id_kualitas = block["choice_map"].get(choice) if choice else None
                nama_kualitas = block["choice_names"].get(choice) if choice else None
            else:
                id_kualitas = block["fixed_h"]
                nama_kualitas = block["fixed_i"]

            d = parse_date(tanggal_raw)
            tanggal = minggu = no_bulan = bulan_nama = periode = no_periode = periode_full = tahun = None
            if d is not None:
                tanggal = d
                minggu = week_of_month(d)
                no_bulan = f"{d.month:02d}"
                bulan_nama = BULAN_ID[d.month - 1]
                periode = f"{BULAN_ID_3[d.month - 1]}-{minggu}"
                no_periode = f"{no_bulan}-{minggu}"
                tahun = d.year
                periode_full = f"{tahun}-{no_periode}"

            row_out = [
                wv["id_kec_desa"], wv["kode_kecamatan"], wv["kecamatan"], wv["kode_desa"], wv["desa"],
                id_kom, nama,
                id_kualitas, nama_kualitas,
                unit,
                harga,
                tanggal,
                minggu, no_bulan, bulan_nama, periode, no_periode,
                asal,
                periode_full, tahun,
            ]
            ws_raw.append(row_out)
            n_raw += 1

            is_blank = (id_kualitas is None and harga in (None, 0, "") and d is None)
            if is_blank:
                n_blank += 1
                continue
            if not wv["valid"]:
                n_dropped_wilayah += 1
                continue
            if harga is None or (isinstance(harga, (int, float)) and harga < min_harga):
                n_dropped_price += 1
                continue
            ws_clean.append(row_out)
            n_clean += 1

    for ws in (ws_raw, ws_clean):
        for row in ws.iter_rows(min_row=2, min_col=12, max_col=12):
            for cell in row:
                if cell.value is not None:
                    cell.number_format = "yyyy-mm-dd"

    buf = io.BytesIO()
    out_wb.save(buf)
    buf.seek(0)

    stats = dict(
        n_rows=n_rows,
        n_raw=n_raw,
        n_clean=n_clean,
        n_blank=n_blank,
        n_dropped_wilayah=n_dropped_wilayah,
        n_dropped_price=n_dropped_price,
    )
    return buf.getvalue(), stats


# ============================================================
# ========================= UI STREAMLIT ======================
# ============================================================

st.set_page_config(page_title="Pipeline Data Harga", page_icon="📊", layout="centered")

st.title("Pipeline Data Harga Kebutuhan Pokok")
st.write(
    "Upload file Excel mentah hasil download aplikasi kuisioner "
    "(harus berisi sheet **1. list_data (ori)**, **kd wilayah**, dan **Kode Kualitas**). "
    "Aplikasi ini otomatis membersihkan kode wilayah, unpivot 20 komoditas, "
    "dan memfilter data valid -- tanpa perlu proses manual sheet demi sheet."
)

uploaded_file = st.file_uploader("Pilih file Excel (.xlsx)", type=["xlsx"])

min_harga = st.number_input(
    "Harga minimum (baris dengan harga di bawah ini dianggap salah ketik dan dibuang di sheet Cleaned)",
    min_value=0, value=1000, step=100,
)

if uploaded_file is not None:
    if st.button("Proses", type="primary"):
        progress_bar = st.progress(0.0, text="Memulai...")

        def progress_cb(frac, nama_komoditas):
            progress_bar.progress(frac, text=f"Memproses: {nama_komoditas}")

        try:
            output_bytes, stats = generate(
                uploaded_file.getvalue(), int(min_harga), progress_cb=progress_cb
            )
        except ValueError as e:
            st.error(str(e))
        except KeyError as e:
            st.error(f"Ada sheet referensi yang formatnya tidak sesuai: {e}")
        else:
            progress_bar.progress(1.0, text="Selesai!")
            st.success("Pipeline selesai dijalankan ✅")

            col1, col2 = st.columns(2)
            with col1:
                st.metric("Total responden terdeteksi", stats["n_rows"])
                st.metric("Baris 'Dashboard (Raw)'", stats["n_raw"])
            with col2:
                st.metric("Baris 'Dashboard (Cleaned)'", stats["n_clean"])
                st.metric(
                    "Total dibuang",
                    stats["n_blank"] + stats["n_dropped_wilayah"] + stats["n_dropped_price"],
                )

            with st.expander("Rincian baris yang dibuang dari sheet Cleaned"):
                st.write(f"- Tanpa jawaban (kualitas/harga/tanggal kosong): **{stats['n_blank']}**")
                st.write(f"- Wilayah gagal lookup (#N/A): **{stats['n_dropped_wilayah']}**")
                st.write(f"- Harga di bawah {min_harga} (dianggap salah ketik): **{stats['n_dropped_price']}**")

            out_name = uploaded_file.name.rsplit(".", 1)[0] + "_hasil.xlsx"
            st.download_button(
                "⬇️ Download hasil (.xlsx)",
                data=output_bytes,
                file_name=out_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
            )
else:
    st.info("Silakan upload file .xlsx untuk mulai.")

st.divider()
st.caption(
    "Catatan: file input cukup berisi sheet '1. list_data (ori)' + 2 sheet referensi "
    "yang tidak berubah ('kd wilayah' dan 'Kode Kualitas'). "
    "Output berisi 2 sheet: 'Dashboard (Raw)' (setara sheet 5 lama) dan "
    "'Dashboard (Cleaned)' (setara sheet 6 lama, sudah difilter)."
)