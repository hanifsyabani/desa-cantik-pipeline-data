"""Excel and parsing helpers for the Desa Cantik pipeline."""

import datetime
import re

from openpyxl.utils import column_index_from_string

from .config import COL_SHIFT_SHEET3_TO_SHEET1, QUALITY_SHEET, WILAYAH_SHEET

CHOICE_RE = re.compile(r"^\s*(\d+)\s*\.")


def shifted(col_letter):
    """Geser huruf kolom dari posisi sheet 3 ke posisi sheet 1."""
    return column_index_from_string(col_letter) + COL_SHIFT_SHEET3_TO_SHEET1


def excel_weeknum(d):
    jan1 = datetime.date(d.year, 1, 1)
    py_wd_jan1 = jan1.weekday()
    excel_wd_jan1 = ((py_wd_jan1 + 1) % 7) + 1
    days_since_jan1 = (d - jan1).days
    return (days_since_jan1 + excel_wd_jan1 - 1) // 7 + 1


def week_of_month(d):
    first_of_month = d.replace(day=1)
    return excel_weeknum(d) - excel_weeknum(first_of_month) + 1


def parse_choice(raw_text):
    if raw_text is None:
        return None

    value = str(raw_text).strip()
    if not value:
        return None

    match = CHOICE_RE.match(value)
    return int(match.group(1)) if match else None


def parse_date(raw):
    if isinstance(raw, datetime.datetime):
        return raw.date()
    if isinstance(raw, datetime.date):
        return raw
    if isinstance(raw, str) and raw.strip():
        try:
            return datetime.datetime.fromisoformat(raw.strip()).date()
        except ValueError:
            return None
    return None


def load_wilayah_lookup(wb):
    """Bangun dict lookup dari sheet 'kd wilayah'."""
    ws = wb[WILAYAH_SHEET]
    kec_name_to_code = {}
    kec_code_to_name = {}
    desa_nospace_to_code = {}
    idbps_to_desaname = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        # Index tuple 0-based: D=3, H=7, I=8, J=9, N=13, O=14, P=15.
        id_bps, desa_h = row[3], row[7]
        if id_bps is not None and desa_h is not None:
            idbps_to_desaname[str(id_bps)] = desa_h

        desa_nospace, kode_desa_bps = row[8], row[9]
        if desa_nospace is not None and kode_desa_bps is not None:
            desa_nospace_to_code[str(desa_nospace)] = kode_desa_bps

        kode_kec_bps_n, kec_o = row[13], row[14]
        if kode_kec_bps_n is not None and kec_o is not None:
            kec_code_to_name[str(kode_kec_bps_n)] = kec_o

        kec_o2, kode_kec_bps_p = row[14], row[15]
        if kec_o2 is not None and kode_kec_bps_p is not None:
            kec_name_to_code[str(kec_o2)] = kode_kec_bps_p

    return kec_name_to_code, kec_code_to_name, desa_nospace_to_code, idbps_to_desaname


def load_quality_names(wb):
    ws = wb[QUALITY_SHEET]
    mapping = {}
    header = [cell.value for cell in ws[1]]
    idx_id = header.index("id_kualitas")
    idx_nama = header.index("nama_kualitas")

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx_id] is None:
            continue
        mapping[int(row[idx_id])] = row[idx_nama]

    return mapping


def clean_wilayah(kec_raw, desa_raw, lookups):
    kec_name_to_code, kec_code_to_name, desa_nospace_to_code, idbps_to_desaname = lookups

    kec_clean = str(kec_raw or "").replace("KECAMATAN ", "").strip()
    desa_step1 = str(desa_raw or "").replace("DESA", "")
    desa_step2 = desa_step1.replace("KELURAHAN", "")
    desa_nospace = desa_step2.replace(" ", "")

    kec_code = kec_name_to_code.get(kec_clean)
    kec_final = kec_code_to_name.get(str(kec_code)) if kec_code is not None else None
    desa_code = desa_nospace_to_code.get(desa_nospace)

    if kec_code is not None and desa_code is not None:
        id_kec_desa = f"{kec_code}{desa_code}"
    else:
        id_kec_desa = "#N/A"

    desa_final = idbps_to_desaname.get(id_kec_desa)
    valid = kec_code is not None and desa_code is not None and desa_final is not None

    return dict(
        id_kec_desa=id_kec_desa if valid else "#N/A",
        kode_kecamatan=kec_code if kec_code is not None else "#N/A",
        kecamatan=kec_final if kec_final else "#N/A",
        kode_desa=desa_code if desa_code is not None else "#N/A",
        desa=desa_final if desa_final else "#N/A",
        valid=valid,
    )
