"""Core transformation pipeline for Desa Cantik workbook data."""

import io

import openpyxl

from .config import (
    BLOCKS,
    BULAN_ID,
    BULAN_ID_3,
    COL_DESA_RAW,
    COL_KECAMATAN_RAW,
    DASHBOARD_CLEANED_SHEET,
    DASHBOARD_RAW_SHEET,
    HEADERS,
    SRC_DATA_START_ROW,
    SRC_SHEET,
)
from .excel_utils import (
    clean_wilayah,
    load_wilayah_lookup,
    parse_choice,
    parse_date,
    shifted,
    week_of_month,
)


def _find_last_data_row(ws):
    last_row = SRC_DATA_START_ROW - 1
    for row_num in range(SRC_DATA_START_ROW, ws.max_row + 1):
        if ws.cell(row=row_num, column=1).value not in (None, ""):
            last_row = row_num
    return last_row


def _format_date_columns(*worksheets):
    for ws in worksheets:
        for row in ws.iter_rows(min_row=2, min_col=12, max_col=12):
            for cell in row:
                if cell.value is not None:
                    cell.number_format = "yyyy-mm-dd"


def build_dashboard_workbook(input_wb, min_harga=1000, progress_cb=None):
    ws1 = input_wb[SRC_SHEET]
    lookups = load_wilayah_lookup(input_wb)

    last_row = _find_last_data_row(ws1)
    n_rows = last_row - SRC_DATA_START_ROW + 1

    out_wb = openpyxl.Workbook()
    ws_raw = out_wb.active
    ws_raw.title = DASHBOARD_RAW_SHEET
    ws_clean = out_wb.create_sheet(DASHBOARD_CLEANED_SHEET)
    ws_raw.append(HEADERS)
    ws_clean.append(HEADERS)

    wilayah_cache = {}
    for row_num in range(SRC_DATA_START_ROW, last_row + 1):
        kec_raw = ws1.cell(row=row_num, column=COL_KECAMATAN_RAW).value
        desa_raw = ws1.cell(row=row_num, column=COL_DESA_RAW).value
        wilayah_cache[row_num] = clean_wilayah(kec_raw, desa_raw, lookups)

    n_raw = 0
    n_clean = 0
    n_blank = 0
    n_dropped_wilayah = 0
    n_dropped_price = 0

    total_blocks = len(BLOCKS)
    for block_index, block in enumerate(BLOCKS):
        if progress_cb:
            progress_cb((block_index + 1) / total_blocks, block["nama"])

        id_kom = block["id_kom"]
        nama = block["nama"]
        unit = block["unit"]
        harga_col = shifted(block["harga_col"])
        tanggal_col = shifted(block["tanggal_col"])
        asal_col = shifted(block["asal_col"])
        kualitas_col = shifted(block["kualitas_col"]) if block["kualitas_col"] else None

        for row_num in range(SRC_DATA_START_ROW, last_row + 1):
            wilayah = wilayah_cache[row_num]

            harga = ws1.cell(row=row_num, column=harga_col).value
            tanggal_raw = ws1.cell(row=row_num, column=tanggal_col).value
            asal = ws1.cell(row=row_num, column=asal_col).value

            if kualitas_col is not None:
                raw_quality = ws1.cell(row=row_num, column=kualitas_col).value
                choice = parse_choice(raw_quality)
                id_kualitas = block["choice_map"].get(choice) if choice else None
                nama_kualitas = block["choice_names"].get(choice) if choice else None
            else:
                id_kualitas = block["fixed_h"]
                nama_kualitas = block["fixed_i"]

            date_value = parse_date(tanggal_raw)
            tanggal = None
            minggu = None
            no_bulan = None
            bulan_nama = None
            periode = None
            no_periode = None
            periode_full = None
            tahun = None

            if date_value is not None:
                tanggal = date_value
                minggu = week_of_month(date_value)
                no_bulan = f"{date_value.month:02d}"
                bulan_nama = BULAN_ID[date_value.month - 1]
                periode = f"{BULAN_ID_3[date_value.month - 1]}-{minggu}"
                no_periode = f"{no_bulan}-{minggu}"
                tahun = date_value.year
                periode_full = f"{tahun}-{no_periode}"

            row_out = [
                wilayah["id_kec_desa"],
                wilayah["kode_kecamatan"],
                wilayah["kecamatan"],
                wilayah["kode_desa"],
                wilayah["desa"],
                id_kom,
                nama,
                id_kualitas,
                nama_kualitas,
                unit,
                harga,
                tanggal,
                minggu,
                no_bulan,
                bulan_nama,
                periode,
                no_periode,
                asal,
                periode_full,
                tahun,
            ]

            ws_raw.append(row_out)
            n_raw += 1

            is_blank = id_kualitas is None and harga in (None, 0, "") and date_value is None
            if is_blank:
                n_blank += 1
                continue
            if not wilayah["valid"]:
                n_dropped_wilayah += 1
                continue
            if harga is None or (isinstance(harga, (int, float)) and harga < min_harga):
                n_dropped_price += 1
                continue

            ws_clean.append(row_out)
            n_clean += 1

    _format_date_columns(ws_raw, ws_clean)

    stats = dict(
        n_rows=n_rows,
        n_raw=n_raw,
        n_clean=n_clean,
        n_blank=n_blank,
        n_dropped_wilayah=n_dropped_wilayah,
        n_dropped_price=n_dropped_price,
        src_data_start_row=SRC_DATA_START_ROW,
        last_row=last_row,
    )
    return out_wb, stats


def generate_bytes(input_bytes, min_harga=1000, progress_cb=None):
    input_wb = openpyxl.load_workbook(io.BytesIO(input_bytes), data_only=True)
    output_wb, stats = build_dashboard_workbook(input_wb, min_harga, progress_cb=progress_cb)

    output = io.BytesIO()
    output_wb.save(output)
    output.seek(0)
    return output.getvalue(), stats


def generate_file(input_path, output_path, min_harga=1000, progress_cb=None):
    input_wb = openpyxl.load_workbook(input_path, data_only=True)
    output_wb, stats = build_dashboard_workbook(input_wb, min_harga, progress_cb=progress_cb)
    output_wb.save(output_path)
    return stats


def generate(input_path, output_path, min_harga=1000):
    return generate_file(input_path, output_path, min_harga)
